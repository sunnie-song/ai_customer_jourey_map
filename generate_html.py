#!/usr/bin/env python3
"""
闲鱼用户旅程 Excel → HTML 生成器
读取「闲鱼用户旅程.xlsx」，生成泳道式交互可视化 HTML

用法：
  python generate_html.py                          # 默认路径
  python generate_html.py -i 路径/文件.xlsx         # 指定 Excel
  python generate_html.py -o 输出.html              # 指定输出路径

Excel 结构：
  Sheet 名包含「卖家」或「买家」
  表头在 row 1，数据从 row 2 开始
  列：阶段区间 | 阶段名称 | 线上行为 | 线下行为 | 触点/渠道 | 用户想法 | 情绪 | 痛点 | 机会点 | 感受说明
"""

import argparse, json, os, sys
from pathlib import Path
from datetime import date

try:
    import openpyxl
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# ============================================================
# Excel 读取
# ============================================================
def read_excel(filepath):
    """读取 Excel 中卖家/买家旅程 Sheet"""
    wb = openpyxl.load_workbook(filepath)
    journeys = []

    for sheet_name in wb.sheetnames:
        if "卖家" in sheet_name:
            journey_type = "seller"
            journey_title = "卖家旅程"
        elif "买家" in sheet_name:
            journey_type = "buyer"
            journey_title = "买家旅程"
        else:
            continue

        ws = wb[sheet_name]

        # 读阶段区间，建映射
        phase_set = []
        for r in range(2, 50):
            phase_val = str(ws.cell(r, 1).value or "").strip()
            if not phase_val:
                break
            if phase_val not in phase_set:
                phase_set.append(phase_val)

        phase_info = {}
        for p in phase_set:
            if "前" in p:
                phase_info[p] = {"label": p, "css": "before", "group": "before"}
            elif "后" in p:
                phase_info[p] = {"label": p, "css": "after", "group": "after"}
            else:
                phase_info[p] = {"label": p, "css": "during", "group": "during"}

        # 读阶段数据（row 2 开始）
        stages = []
        for r in range(2, 50):
            phase_val = str(ws.cell(r, 1).value or "").strip()
            if not phase_val:
                break

            name = str(ws.cell(r, 2).value or "").strip()
            online = str(ws.cell(r, 3).value or "—").strip()
            offline = str(ws.cell(r, 4).value or "—").strip()
            touchpoint = str(ws.cell(r, 5).value or "").strip()
            thought = str(ws.cell(r, 6).value or "").strip()
            emotion_raw = str(ws.cell(r, 7).value or "中").strip()
            pain_raw = str(ws.cell(r, 8).value or "").strip()
            opp_raw = str(ws.cell(r, 9).value or "").strip()
            feeling = str(ws.cell(r, 10).value or "").strip()

            # 情绪映射
            if "高" in emotion_raw:
                emotion, emoji = "high", "😁"
            elif "低" in emotion_raw:
                emotion, emoji = "low", "😞"
            else:
                emotion, emoji = "mid", "😐"

            # 痛点 / 机会点拆分
            pains = [p.strip() for p in pain_raw.replace("\n", "|").split("|") if p.strip()]
            opps = [o.strip() for o in opp_raw.replace("\n", "|").split("|") if o.strip()]

            # 合并行为
            actions = []
            if online and online != "—": actions.append(online)
            if offline and offline != "—": actions.append(offline)
            action_text = " / ".join(actions) if actions else "—"

            stages.append({
                "name": name,
                "phase_label": phase_val,
                "phase_group": phase_info[phase_val]["group"],
                "phase_css": phase_info[phase_val]["css"],
                "emotion": emotion, "emoji": emoji,
                "action": action_text,
                "touchpoint": touchpoint,
                "thought": thought,
                "pains": pains, "opps": opps,
                "feeling": feeling,
            })

        # 阶段分组
        phase_groups = []
        seen = set()
        for s in stages:
            key = s["phase_group"]
            if key not in seen:
                seen.add(key)
                group_stages = [st["name"] for st in stages if st["phase_group"] == key]
                name_map = {"before": "使用前", "during": "使用中", "after": "使用后"}
                phase_groups.append({
                    "name": name_map.get(key, key),
                    "css": key,
                    "stages": group_stages,
                    "count": len(group_stages),
                })

        journeys.append({
            "id": journey_type,
            "title": journey_title,
            "phase_groups": phase_groups,
            "stages": stages,
        })

    return journeys


# ============================================================
# CSS
# ============================================================
CSS = r"""
:root {
  --bg: #F5F4F1;
  --lane-bg: #FFFFFF;
  --lane-alt: #FAFAF8;
  --text: #1A1A1A;
  --text-secondary: #5F5E5A;
  --text-muted: #8B8A85;
  --border: #DEDCD5;
  --border-light: #EBE9E3;
  --purple: #534AB7;
  --orange: #D85A30;
  --orange-light: #FAECE7;
  --teal: #0F6E56;
  --teal-light: #E1F5EE;
  --amber: #BA7517;
  --amber-light: #FAEEDA;
  --red: #E24B4A;
  --red-light: #FCEBEB;
  --blue: #378ADD;
  --blue-light: #E6F1FB;
  --green: #3B6D11;
  --green-light: #EAF3DE;
}
* { margin: 0; padding: 0; box-sizing: border-box; }
body {
  font-family: -apple-system, BlinkMacSystemFont, "PingFang SC", "Hiragino Sans GB", "Microsoft YaHei", sans-serif;
  background: var(--bg); color: var(--text);
  -webkit-font-smoothing: antialiased; overflow-x: hidden;
}
.top-bar { background: #1A1A1A; color: #FFF; padding: 16px 32px; display: flex; align-items: center; justify-content: space-between; }
.top-bar h1 { font-size: 15px; font-weight: 600; }
.top-bar .meta { font-size: 11px; color: #999; }
.tab-bar { display: flex; gap: 0; background: #FFF; border-bottom: 1px solid var(--border); padding: 0 32px; }
.tab-btn { padding: 12px 24px; font-size: 13px; font-weight: 500; cursor: pointer; border: none; background: none; color: var(--text-secondary); border-bottom: 2px solid transparent; transition: all 0.15s; }
.tab-btn:hover { color: var(--text); }
.tab-btn.active { color: var(--purple); border-bottom-color: var(--purple); }
.map-wrapper { margin: 24px 32px; background: #FFF; border-radius: 12px; border: 1px solid var(--border); overflow: hidden; box-shadow: 0 1px 4px rgba(0,0,0,0.04); }
.map-scroll { overflow-x: auto; overflow-y: hidden; }
.map-scroll::-webkit-scrollbar { height: 8px; }
.map-scroll::-webkit-scrollbar-track { background: var(--bg); border-radius: 4px; }
.map-scroll::-webkit-scrollbar-thumb { background: var(--border); border-radius: 4px; }
.phase-row { display: flex; border-bottom: 1px solid var(--border); position: sticky; top: 0; z-index: 10; background: #FFF; }
.phase-label-cell { width: 120px; min-width: 120px; flex-shrink: 0; border-right: 1px solid var(--border); display: flex; align-items: center; justify-content: center; font-size: 11px; font-weight: 600; color: var(--text-muted); letter-spacing: 1px; background: #FAFAF8; }
.phase-group { display: flex; flex-direction: column; flex-shrink: 0; }
.phase-header { padding: 12px 0; text-align: center; font-size: 12px; font-weight: 600; border-bottom: 1px solid var(--border-light); }
.phase-header.before { background: var(--orange-light); color: var(--orange); }
.phase-header.during { background: var(--amber-light); color: var(--amber); }
.phase-header.after { background: var(--teal-light); color: var(--teal); }
.phase-stages { display: flex; }
.stage-col { width: 160px; min-width: 160px; flex-shrink: 0; text-align: center; padding: 10px 8px; border-right: 1px solid var(--border-light); font-size: 11px; font-weight: 600; color: var(--text-secondary); line-height: 1.4; }
.stage-col:last-child { border-right: none; }
.lane-row { display: flex; border-bottom: 1px solid var(--border-light); }
.lane-row:last-child { border-bottom: none; }
.lane-label { width: 120px; min-width: 120px; flex-shrink: 0; border-right: 1px solid var(--border); display: flex; align-items: center; padding: 12px 14px; font-size: 11px; font-weight: 600; color: var(--text-muted); letter-spacing: 0.5px; background: #FAFAF8; }
.lane-label .lane-icon { margin-right: 6px; font-size: 14px; }
.lane-content { display: flex; }
.cell { width: 160px; min-width: 160px; flex-shrink: 0; padding: 12px 10px; border-right: 1px solid var(--border-light); font-size: 12px; line-height: 1.5; color: var(--text-secondary); display: flex; flex-direction: column; justify-content: center; }
.cell:last-child { border-right: none; }
.lane-row.alt .lane-label { background: #FFF; }
.lane-row.alt .cell { background: var(--lane-alt); }
.emotion-curve-cell { width: 160px; min-width: 160px; flex-shrink: 0; border-right: 1px solid var(--border-light); }
.emotion-curve-cell:last-child { border-right: none; }
.pain-tag { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; margin: 2px 0; line-height: 1.5; background: var(--red-light); color: #791F1F; }
.opp-tag { display: inline-block; padding: 2px 8px; border-radius: 4px; font-size: 11px; margin: 2px 0; line-height: 1.5; background: var(--blue-light); color: #0C447C; }
.quote-text { font-style: italic; color: var(--text-muted); position: relative; padding-left: 8px; border-left: 2px solid var(--purple); }
.legend-bar { display: flex; gap: 24px; padding: 16px 32px; font-size: 12px; color: var(--text-muted); align-items: center; }
.legend-item { display: flex; align-items: center; gap: 6px; }
.legend-swatch { width: 12px; height: 12px; border-radius: 3px; }
.footer { padding: 8px 32px 24px; font-size: 11px; color: var(--text-muted); text-align: right; }
@media (max-width: 768px) {
  .map-wrapper { margin: 12px 8px; border-radius: 8px; }
  .phase-label-cell, .lane-label { width: 80px; min-width: 80px; font-size: 10px; }
  .stage-col, .cell { width: 130px; min-width: 130px; font-size: 11px; }
  .tab-bar, .top-bar, .legend-bar { padding-left: 12px; padding-right: 12px; }
}
"""


# ============================================================
# JavaScript 模板
# ============================================================
JS_TEMPLATE = r"""
const JOURNEYS = __JOURNEYS_JSON__;
let currentTab = '__DEFAULT_TAB__';

function renderMap(journey) {
  const container = document.getElementById('map-scroll');
  let html = '';

  // 阶段头部行
  html += '<div class="phase-row">';
  html += '<div class="phase-label-cell">阶段</div>';
  for (const pg of journey.phase_groups) {
    html += '<div class="phase-group">';
    html += '<div class="phase-header ' + pg.css + '" style="width:' + (pg.count * 160) + 'px">' + pg.name + '</div>';
    html += '<div class="phase-stages">';
    for (const sn of pg.stages) {
      const full = journey.stages.find(s => s.name === sn);
      html += '<div class="stage-col">' + (full ? full.emoji + ' ' : '') + sn + '</div>';
    }
    html += '</div></div>';
  }
  html += '</div>';

  // 泳道 1: 用户行为
  html += '<div class="lane-row">';
  html += '<div class="lane-label"><span class="lane-icon">👤</span>用户行为</div>';
  html += '<div class="lane-content">';
  for (const s of journey.stages) {
    html += '<div class="cell">' + esc(s.action) + '</div>';
  }
  html += '</div></div>';

  // 泳道 2: 触点/渠道
  html += '<div class="lane-row alt">';
  html += '<div class="lane-label"><span class="lane-icon">📍</span>触点 / 渠道</div>';
  html += '<div class="lane-content">';
  for (const s of journey.stages) {
    html += '<div class="cell">' + esc(s.touchpoint) + '</div>';
  }
  html += '</div></div>';

  // 泳道 3: 用户想法
  html += '<div class="lane-row">';
  html += '<div class="lane-label"><span class="lane-icon">💬</span>用户想法</div>';
  html += '<div class="lane-content">';
  for (const s of journey.stages) {
    html += '<div class="cell"><div class="quote-text">"' + esc(s.thought) + '"</div></div>';
  }
  html += '</div></div>';

  // 泳道 4: 情绪曲线
  html += '<div class="lane-row alt">';
  html += '<div class="lane-label"><span class="lane-icon">📊</span>情绪曲线</div>';
  html += '<div class="lane-content">';
  for (const s of journey.stages) {
    const label = s.emotion === 'high' ? '满意' : s.emotion === 'mid' ? '一般' : '不满';
    const h = s.emotion === 'high' ? 52 : s.emotion === 'mid' ? 32 : 16;
    const barColor = s.emotion === 'high' ? '#3B6D11' : s.emotion === 'mid' ? '#BA7517' : '#E24B4A';
    html += '<div class="emotion-curve-cell" style="display:flex;align-items:flex-end;justify-content:center;padding-bottom:6px;">';
    html += '<div style="text-align:center;">';
    html += '<div style="font-size:22px;margin-bottom:4px;">' + s.emoji + '</div>';
    html += '<div style="width:8px;height:' + h + 'px;border-radius:4px;background:' + barColor + ';margin:0 auto 4px;"></div>';
    html += '<div style="font-size:10px;color:var(--text-muted)">' + label + '</div>';
    html += '</div></div>';
  }
  html += '</div></div>';

  // 泳道 5: 痛点
  html += '<div class="lane-row">';
  html += '<div class="lane-label"><span class="lane-icon">🔴</span>痛点</div>';
  html += '<div class="lane-content">';
  for (const s of journey.stages) {
    html += '<div class="cell">';
    for (const p of s.pains) {
      html += '<span class="pain-tag">' + esc(p) + '</span>';
    }
    html += '</div>';
  }
  html += '</div></div>';

  // 泳道 6: 机会点
  html += '<div class="lane-row alt">';
  html += '<div class="lane-label"><span class="lane-icon">🔵</span>机会点</div>';
  html += '<div class="lane-content">';
  for (const s of journey.stages) {
    html += '<div class="cell">';
    for (const o of s.opps) {
      html += '<span class="opp-tag">' + esc(o) + '</span>';
    }
    html += '</div>';
  }
  html += '</div></div>';

  container.innerHTML = html;
}

function esc(s) { return String(s || '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;'); }

function switchJourney(id) {
  currentTab = id;
  document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
  document.getElementById('tab-' + id).classList.add('active');
  const j = JOURNEYS.find(j => j.id === id);
  if (j) renderMap(j);
}

document.addEventListener('DOMContentLoaded', function() {
  switchJourney(currentTab);
});
"""


# ============================================================
# HTML 构建
# ============================================================
def build_html(journeys, default_tab="seller"):
    journeys_json = json.dumps(journeys, ensure_ascii=False, indent=2)

    tab_buttons = ""
    for j in journeys:
        icon = "🏪" if j["id"] == "seller" else "🛒"
        tab_buttons += f'<button class="tab-btn" onclick="switchJourney(\'{j["id"]}\')" id="tab-{j["id"]}">{icon} {j["title"]}</button>\n'

    js_code = JS_TEMPLATE.replace("__JOURNEYS_JSON__", journeys_json).replace("__DEFAULT_TAB__", default_tab)

    today = date.today()

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>闲鱼用户旅程图</title>
<style>{CSS}</style>
</head>
<body>
<div class="top-bar">
  <h1>闲鱼用户旅程图</h1>
  <div class="meta">由 Excel 自动生成 | 卖家 / 买家双旅程</div>
</div>
<div class="tab-bar">{tab_buttons}</div>
<div class="legend-bar">
  <span>情绪图例：</span>
  <div class="legend-item"><div class="legend-swatch" style="background:#3B6D11"></div>高满意度</div>
  <div class="legend-item"><div class="legend-swatch" style="background:#BA7517"></div>中满意度</div>
  <div class="legend-item"><div class="legend-swatch" style="background:#E24B4A"></div>低满意度</div>
  <span style="margin-left:16px">标注图例：</span>
  <div class="legend-item"><div class="legend-swatch" style="background:#FCEBEB;border:1px solid #F09595"></div>🔴 痛点</div>
  <div class="legend-item"><div class="legend-swatch" style="background:#E6F1FB;border:1px solid #85B7EB"></div>🔵 机会点</div>
</div>
<div class="map-wrapper"><div class="map-scroll" id="map-scroll"></div></div>
<div class="footer">闲鱼用户旅程图 · 数据来源：闲鱼用户旅程.xlsx · 生成日期 {today}</div>
<script>{js_code}</script>
</body>
</html>"""
    return html


# ============================================================
# 主入口
# ============================================================
def main():
    parser = argparse.ArgumentParser(description="闲鱼用户旅程 Excel → HTML 生成器")
    parser.add_argument("-i", "--input", default=None, help="Excel 文件路径")
    parser.add_argument("-o", "--output", default=None, help="输出 HTML 路径")
    args = parser.parse_args()

    script_dir = Path(__file__).parent
    default_excel = script_dir / "闲鱼用户旅程.xlsx"
    default_html = script_dir / "journey_map.html"

    excel_path = Path(args.input) if args.input else default_excel
    html_path = Path(args.output) if args.output else default_html

    if not excel_path.exists():
        print(f"错误：找不到 Excel 文件 {excel_path}")
        print(f"用法：python {Path(__file__).name} -i 路径/文件.xlsx")
        sys.exit(1)

    print(f"📄 读取 Excel: {excel_path}")
    journeys = read_excel(str(excel_path))

    if not journeys:
        print("错误：未找到有效的旅程 Sheet（Sheet 名需包含「卖家」或「买家」）")
        sys.exit(1)

    print(f"找到 {len(journeys)} 个旅程:")
    for j in journeys:
        print(f"  - {j['id']}: {j['title']} ({len(j['stages'])} 个阶段)")

    default_tab = journeys[0]["id"]
    html = build_html(journeys, default_tab)

    html_path.parent.mkdir(parents=True, exist_ok=True)
    html_path.write_text(html, encoding="utf-8")
    print(f"✨ 生成 HTML: {html_path} ({len(html)} 字节)")
    print(f"👉 用浏览器打开 {html_path} 即可查看")


if __name__ == "__main__":
    main()
