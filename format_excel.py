#!/usr/bin/env python3
"""格式化闲鱼用户旅程 Excel — 参考迪卡侬文件样式"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from copy import copy
import os

filepath = os.path.join(os.path.dirname(__file__), '闲鱼用户旅程.xlsx')
wb = openpyxl.load_workbook(filepath)

# ---- Border style (thin, light gray) ----
thin_border = Border(
    left=Side(style='thin', color='D0CDC4'),
    right=Side(style='thin', color='D0CDC4'),
    top=Side(style='thin', color='D0CDC4'),
    bottom=Side(style='thin', color='D0CDC4'),
)

# ---- Phase colors (matching the Decathlon reference) ----
phase_fills = {
    '使用前': PatternFill(start_color='FAECE7', end_color='FAECE7', fill_type='solid'),  # orange-light
    '使用中': PatternFill(start_color='FAEEDA', end_color='FAEEDA', fill_type='solid'),  # amber-light
    '使用后': PatternFill(start_color='E1F5EE', end_color='E1F5EE', fill_type='solid'),  # teal-light
}

phase_font_colors = {
    '使用前': 'D85A30',
    '使用中': 'BA7517',
    '使用后': '0F6E56',
}

# ---- Column widths ----
col_widths = {
    'A': 18,   # 阶段区间
    'B': 18,   # 阶段名称
    'C': 34,   # 线上行为
    'D': 24,   # 线下行为
    'E': 22,   # 触点/渠道
    'F': 36,   # 用户想法
    'G': 8,    # 情绪
    'H': 30,   # 痛点
    'I': 32,   # 机会点
    'J': 24,   # 感受说明
}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    # -- Column widths --
    for col_letter, width in col_widths.items():
        if col_letter in ws.column_dimensions:
            ws.column_dimensions[col_letter].width = width
        else:
            ws.column_dimensions[col_letter] = openpyxl.worksheet.dimensions.ColumnDimension(ws, index=openpyxl.utils.column_index_from_string(col_letter), width=width)

    # -- Row 1: Header --
    header_fill = PatternFill(start_color='2C2C2A', end_color='2C2C2A', fill_type='solid')
    header_font = Font(name='Inter', size=11, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 36

    for c in range(1, max_col + 1):
        cell = ws.cell(1, c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # -- Data rows (row 2 onwards) --
    data_font = Font(name='Inter', size=10, color='1A1A1A')
    data_font_bold = Font(name='Inter', size=10, bold=True, color='1A1A1A')
    data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for r in range(2, max_row + 1):
        ws.row_dimensions[r].height = 80  # taller for wrapped content

        # Get phase value from col A
        phase_val = str(ws.cell(r, 1).value or '').strip()
        phase_fill = phase_fills.get(phase_val)
        phase_color = phase_font_colors.get(phase_val)

        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.border = thin_border
            cell.alignment = data_align
            cell.font = data_font

            # Phase column (A): colored bg, bold, colored text
            if c == 1 and phase_fill:
                cell.fill = phase_fill
                cell.font = Font(name='Inter', size=10, bold=True, color=phase_color)
                cell.alignment = center_align
            # Stage name column (B): bold
            elif c == 2:
                cell.font = data_font_bold
                cell.alignment = center_align
            # Emotion column (G): centered
            elif c == 7:
                cell.alignment = center_align
                emo_val = str(cell.value or '').strip()
                if emo_val == '高':
                    cell.font = Font(name='Inter', size=11, bold=True, color='3B6D11')
                elif emo_val == '低':
                    cell.font = Font(name='Inter', size=11, bold=True, color='E24B4A')
                else:
                    cell.font = Font(name='Inter', size=11, bold=True, color='BA7517')
            # Other columns: normal
            else:
                cell.font = data_font

output = os.path.join(os.path.dirname(__file__), '闲鱼用户旅程.xlsx')
wb.save(output)
print(f'✅ 样式优化完成: {output}')
